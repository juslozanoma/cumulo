#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Cúmulo Knowledge Base Generator v2.0
=====================================
Transforma una colección de documentos (PDF, DOCX, MD, TXT, XLSX, CHATS)
en una base de conocimiento estructurada, enriquecida y optimizada para
consultas mediante modelos de lenguaje.

Novedades v2.0:
    - Soporte completo para Excel (.xlsx, .xls) con todas las hojas
    - Soporte nativo para chats de WhatsApp (.txt exportados)
    - Chats se procesan línea por línea con metadatos de participante
    - Chats se exportan como archivos JSON separados para acceso directo
    - Mejor chunking que preserva contexto de conversaciones

Flujo:
    documentos/  →  docs/kb.json  +  docs/kb-metadata.json  +  docs/chats/

Dependencias:
    - pypdf
    - python-docx
    - openpyxl (o pandas)

Ejecutar:
    python procesar.py

Autor:  Generado para Cúmulo
Fecha:  2026
"""

import os
import json
import re
import hashlib
import unicodedata
from datetime import datetime
from pathlib import Path
from collections import Counter, defaultdict
from typing import List, Dict, Any, Optional, Tuple, Set

# ---------------------------------------------------------------------------
# DEPENDENCIAS EXTERNAS
# ---------------------------------------------------------------------------
try:
    import pypdf
except ImportError:
    print("❌ Falta pypdf. Ejecuta: pip install pypdf")
    exit(1)

try:
    from docx import Document
except ImportError:
    print("❌ Falta python-docx. Ejecuta: pip install python-docx")
    exit(1)

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("⚠️  openpyxl no disponible. Soporte Excel limitado.")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    print("⚠️  pandas no disponible. Algunas funciones de Excel usarán openpyxl.")


# ---------------------------------------------------------------------------
# CONSTANTES Y CONFIGURACIÓN
# ---------------------------------------------------------------------------

# Carpetas de entrada y salida
INPUT_DIR = Path("documentos")
OUTPUT_DIR = Path("docs")
CHATS_DIR = OUTPUT_DIR / "chats"  # ← NUEVO: carpeta para chats exportados

# Tamaños y límites
CHUNK_MAX_WORDS = 600          # Palabras máximas por chunk
CHUNK_MAX_CHARS = 4000         # Caracteres máximos por chunk (seguridad)

# Configuración de chats
CHAT_CHUNK_MESSAGES = 50       # Mensajes por chunk de chat
CHAT_EXPORT_LINES = True       # Exportar chats como archivos JSON separados

MESES_ES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
}

# Tipos de documento detectables
DOCUMENT_TYPES = [
    'acta', 'manual', 'reglamento', 'protocolo', 'relatoria',
    'carta', 'cronograma', 'proyecto', 'informe', 'idea', 'formato',
    'chat', 'excel', 'otro'
]

# Comités detectables
COMMITTEES = [
    'tesoreria', 'astrolectura', 'astroescritura', 'etica', 'charlas',
    'administrativo', 'representacion', 'cumulo nacional', 'cumulo bogota',
    'observatorio', 'universidad', 'general', 'vip', 'inventario',
    'astrochaza', 'viaticos', 'protocolos', 'astronomia',
]

# Alcances detectables
SCOPES = [
    'cumulo bogota', 'cumulo nacional', 'general', 'observatorio',
    'universidad', 'otro',
]

# Estados de documento
STATUSES = ['vigente', 'historico', 'borrador', 'propuesta', 'desconocido']

# Cargos detectables
ROLES = [
    'representante', 'coordinador', 'fundador', 'tesorero', 'relator',
    'moderador', 'asistente', 'presidente', 'vicepresidente', 'secretario',
    'director', 'subdirector', 'miembro', 'invitado', 'ponente',
    'organizador', 'voluntario', 'asesor', 'delegado',
]

# Secciones detectables (en orden de prioridad)
SECTION_HEADERS = [
    'contexto', 'objetivos', 'objetivo', 'asistentes', 'participantes',
    'aprendizajes', 'acuerdos', 'resultados', 'observaciones',
    'tareas realizadas', 'tareas pendientes', 'tareas asignadas',
    'cronograma', 'fechas importantes', 'firmas', 'firma',
    'desarrollo', 'conclusiones', 'recomendaciones', 'anexos',
    'introduccion', 'introducción', 'antecedentes', 'justificacion',
    'justificación', 'metodologia', 'metodología', 'alcance',
    'resumen', 'abstract', 'tabla de contenido', 'contenido',
    'agenda', 'orden del dia', 'orden del día', 'acta',
    'bitacora', 'bitácora', 'notas', 'comentarios',
]

# Palabras vacías (stopwords) en español para keywords
STOPWORDS_ES = {
    'de', 'la', 'que', 'el', 'en', 'y', 'a', 'los', 'del', 'se', 'las',
    'por', 'un', 'para', 'con', 'no', 'una', 'su', 'al', 'lo', 'más',
    'pero', 'sus', 'le', 'ya', 'o', 'este', 'sí', 'porque', 'esta',
    'entre', 'cuando', 'muy', 'sin', 'sobre', 'también', 'me', 'hasta',
    'hay', 'donde', 'quien', 'desde', 'todo', 'nos', 'durante', 'todos',
    'uno', 'les', 'ni', 'contra', 'otros', 'ese', 'eso', 'ante', 'ellos',
    'e', 'esto', 'mí', 'antes', 'algunos', 'qué', 'unos', 'yo', 'otro',
    'otras', 'otra', 'él', 'tanto', 'esa', 'estos', 'mucho', 'quienes',
    'nada', 'muchos', 'cual', 'poco', 'ella', 'estar', 'estas', 'algunas',
    'algo', 'nosotros', 'mi', 'mis', 'tú', 'te', 'ti', 'tu', 'tus',
    'ellas', 'nosotras', 'vosotros', 'vosotras', 'os', 'mío', 'mía',
    'míos', 'mías', 'tuyo', 'tuya', 'tuyos', 'tuyas', 'suyo', 'suya',
    'suyos', 'suyas', 'nuestro', 'nuestra', 'nuestros', 'nuestras',
    'vuestro', 'vuestra', 'vuestros', 'vuestras', 'esos', 'esas', 'estoy',
    'estás', 'está', 'estamos', 'estáis', 'están', 'esté', 'estés',
    'estemos', 'estéis', 'estén', 'estaré', 'estarás', 'estará',
    'estaremos', 'estaréis', 'estarán', 'estaría', 'estarías',
    'estaríamos', 'estaríais', 'estarían', 'estaba', 'estabas',
    'estábamos', 'estabais', 'estaban', 'estuve', 'estuviste', 'estuvo',
    'estuvimos', 'estuvisteis', 'estuvieron', 'estuviera', 'estuvieras',
    'estuviéramos', 'estuvierais', 'estuvieran', 'estuviese', 'estuvieses',
    'estuviésemos', 'estuvieseis', 'estuviesen', 'estando', 'estado',
    'estada', 'estados', 'estadas', 'estad', 'he', 'has', 'ha', 'hemos',
    'habéis', 'han', 'haya', 'hayas', 'hayamos', 'hayáis', 'hayan',
    'habré', 'habrás', 'habrá', 'habremos', 'habréis', 'habrán',
    'habría', 'habrías', 'habríamos', 'habríais', 'habrían', 'había',
    'habías', 'habíamos', 'habíais', 'habían', 'hube', 'hubiste',
    'hubo', 'hubimos', 'hubisteis', 'hubieron', 'hubiera', 'hubieras',
    'hubiéramos', 'hubierais', 'hubieran', 'hubiese', 'hubieses',
    'hubiésemos', 'hubieseis', 'hubiesen', 'habiendo', 'habido', 'habida',
    'habidos', 'habidas', 'soy', 'eres', 'es', 'somos', 'sois', 'son',
    'sea', 'seas', 'seamos', 'seáis', 'sean', 'seré', 'serás', 'será',
    'seremos', 'seréis', 'serán', 'sería', 'serías', 'seríamos',
    'seríais', 'serían', 'era', 'eras', 'éramos', 'erais', 'eran', 'fui',
    'fuiste', 'fue', 'fuimos', 'fuisteis', 'fueron', 'fuera', 'fueras',
    'fuéramos', 'fuerais', 'fueran', 'fuese', 'fueses', 'fuésemos',
    'fueseis', 'fuesen', 'siendo', 'sido', 'tengo', 'tienes', 'tiene',
    'tenemos', 'tenéis', 'tienen', 'tenga', 'tengas', 'tengamos',
    'tengáis', 'tengan', 'tendré', 'tendrás', 'tendrá', 'tendremos',
    'tendréis', 'tendrán', 'tendría', 'tendrías', 'tendríamos',
    'tendríais', 'tendrían', 'tenía', 'tenías', 'teníamos', 'teníais',
    'tenían', 'tuve', 'tuviste', 'tuvo', 'tuvimos', 'tuvisteis',
    'tuvieron', 'tuviera', 'tuvieras', 'tuviéramos', 'tuvierais',
    'tuvieran', 'tuviese', 'tuvieses', 'tuviésemos', 'tuvieseis',
    'tuviesen', 'teniendo', 'tenido', 'tenida', 'tenidos', 'tenidas',
    'tened',
}


# ===========================================================================
# NUEVO: FUNCIONES DE DETECCIÓN Y PARSEO DE CHATS DE WHATSAPP
# ===========================================================================

def is_whatsapp_chat(text: str, filename: str) -> bool:
    """
    Detecta si un archivo de texto es un chat de WhatsApp exportado.

    Heurísticas:
        - Patrones de fecha/hora típicos de WhatsApp
        - Nombres de archivo comunes (chat, whatsapp, conversación)
        - Presencia de líneas con formato "DD/MM/YY, HH:MM - Nombre: mensaje"

    Args:
        text: Contenido del archivo.
        filename: Nombre del archivo.

    Returns:
        True si parece ser un chat de WhatsApp.
    """
    # Patrones de fecha/hora de WhatsApp (varios formatos)
    whatsapp_patterns = [
        r'^\d{1,2}/\d{1,2}/\d{2,4},?\s+\d{1,2}:\d{2}\s*[APap]?[Mm]?\s*[-–]\s*',
        r'^\[\d{1,2}/\d{1,2}/\d{2,4},?\s+\d{1,2}:\d{2}\s*[APap]?[Mm]?\]\s*',
        r'^\d{1,2}\.\d{1,2}\.\d{2,4},?\s+\d{1,2}:\d{2}\s*[-–]\s*',
        r'^\d{1,2}-\d{1,2}-\d{2,4},?\s+\d{1,2}:\d{2}\s*[-–]\s*',
    ]

    lines = text.split('\n')
    match_count = 0
    for line in lines[:50]:
        for pattern in whatsapp_patterns:
            if re.match(pattern, line.strip()):
                match_count += 1
                break

    if match_count >= 5:
        return True

    fname_lower = filename.lower()
    chat_keywords = ['chat', 'whatsapp', 'conversacion', 'conversación', 'wa_']
    if any(kw in fname_lower for kw in chat_keywords):
        if match_count >= 2:
            return True

    return False


def parse_whatsapp_chat(text: str) -> List[Dict[str, Any]]:
    """
    Parsea un chat de WhatsApp exportado en una lista de mensajes estructurados.

    Soporta múltiples formatos de fecha/hora.

    Args:
        text: Contenido del chat.

    Returns:
        Lista de mensajes: {date, time, sender, message, type, media}
    """
    messages = []

    patterns = [
        (r'^(\d{1,2}/\d{1,2}/\d{2,4}),?\s+(\d{1,2}:\d{2}(?::\d{2})?)\s*(?:[APap][Mm])?\s*[-–]\s*(.*?):(.*)$', 'slash'),
        (r'^\[(\d{1,2}/\d{1,2}/\d{2,4}),?\s+(\d{1,2}:\d{2}(?::\d{2})?)\s*(?:[APap][Mm])?\]\s*(.*?):(.*)$', 'bracket'),
        (r'^(\d{1,2}\.\d{1,2}\.\d{2,4}),?\s+(\d{1,2}:\d{2}(?::\d{2})?)\s*(?:[APap][Mm])?\s*[-–]\s*(.*?):(.*)$', 'dot'),
        (r'^(\d{1,2}-\d{1,2}-\d{2,4}),?\s+(\d{1,2}:\d{2}(?::\d{2})?)\s*(?:[APap][Mm])?\s*[-–]\s*(.*?):(.*)$', 'dash'),
    ]

    lines = text.split('\n')
    current_message = None

    for line in lines:
        line = line.strip()
        if not line:
            continue

        matched = False
        for pattern, fmt in patterns:
            m = re.match(pattern, line)
            if m:
                if current_message:
                    messages.append(current_message)

                date_str, time_str, sender, msg_text = m.groups()
                msg_text = msg_text.strip()

                msg_type = 'text'
                media_info = None

                if '<Multimedia omitido>' in msg_text or 'omitted' in msg_text.lower():
                    msg_type = 'media_omitted'
                    media_info = {'type': 'unknown', 'note': 'Multimedia omitido'}
                elif msg_text.startswith('http://') or msg_text.startswith('https://'):
                    msg_type = 'link'
                elif msg_text in ['‎', '\u200e', '']:
                    msg_type = 'system'

                current_message = {
                    'date': date_str.strip(),
                    'time': time_str.strip(),
                    'sender': sender.strip(),
                    'message': msg_text,
                    'type': msg_type,
                    'media': media_info,
                    'format': fmt,
                }
                matched = True
                break

        if not matched and current_message:
            current_message['message'] += '\n' + line

    if current_message:
        messages.append(current_message)

    return messages


def chat_to_text(messages: List[Dict[str, Any]]) -> str:
    """Convierte mensajes parseados a texto plano para indexación."""
    lines = []
    for msg in messages:
        lines.append(f"[{msg['date']} {msg['time']}] {msg['sender']}: {msg['message']}")
    return "\n".join(lines)


def export_chat_json(messages: List[Dict[str, Any]], doc_id: str, filename: str) -> Path:
    """
    Exporta un chat como archivo JSON independiente en docs/chats/.

    Args:
        messages: Lista de mensajes parseados.
        doc_id: ID del documento.
        filename: Nombre original del archivo.

    Returns:
        Ruta al archivo JSON exportado.
    """
    CHATS_DIR.mkdir(parents=True, exist_ok=True)

    safe_name = re.sub(r'[^\w\-\.]', '_', filename)
    safe_name = safe_name.replace('.txt', '.json')
    if not safe_name.endswith('.json'):
        safe_name += '.json'

    output_path = CHATS_DIR / safe_name

    chat_export = {
        'document_id': doc_id,
        'source_file': filename,
        'exported_at': datetime.now().isoformat(),
        'total_messages': len(messages),
        'participants': sorted(set(m['sender'] for m in messages)),
        'date_range': {
            'first': messages[0]['date'] if messages else None,
            'last': messages[-1]['date'] if messages else None,
        },
        'messages': messages,
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(chat_export, f, ensure_ascii=False, indent=2)

    return output_path


def chunk_chat_messages(messages: List[Dict[str, Any]], doc_id: str) -> List[Dict[str, Any]]:
    """
    Divide mensajes de chat en chunks preservando contexto conversacional.

    Args:
        messages: Lista de mensajes parseados.
        doc_id: ID del documento.

    Returns:
        Lista de chunks con metadata de chat.
    """
    chunks = []
    total = len(messages)

    for i in range(0, total, CHAT_CHUNK_MESSAGES):
        batch = messages[i:i + CHAT_CHUNK_MESSAGES]

        lines = []
        participants_in_chunk = set()
        dates_in_chunk = set()

        for msg in batch:
            lines.append(f"[{msg['date']} {msg['time']}] {msg['sender']}: {msg['message']}")
            participants_in_chunk.add(msg['sender'])
            dates_in_chunk.add(msg['date'])

        chunk_text = "\n".join(lines)
        words = len(chunk_text.split())

        chunks.append({
            'id': f"chunk_{len(chunks):05d}",
            'document_id': doc_id,
            'section': f"chat_messages_{i+1}_to_{min(i+CHAT_CHUNK_MESSAGES, total)}",
            'text': chunk_text,
            'tokens': words,
            'chars': len(chunk_text),
            'chat_meta': {
                'message_start': i + 1,
                'message_end': min(i + CHAT_CHUNK_MESSAGES, total),
                'total_messages': len(batch),
                'participants': sorted(participants_in_chunk),
                'dates': sorted(dates_in_chunk),
            }
        })

    return chunks


# ===========================================================================
# NUEVO: FUNCIONES DE EXTRACCIÓN DE EXCEL
# ===========================================================================

def extract_text_from_excel(path: Path) -> Tuple[str, int, List[Dict[str, Any]]]:
    """
    Extrae texto de todas las hojas de un archivo Excel.

    Usa openpyxl para preservar el formato y estructura de las tablas.
    Convierte cada hoja a texto Markdown-like para mejor legibilidad.

    Args:
        path: Ruta al archivo Excel (.xlsx o .xls).

    Returns:
        Tupla (texto_extraído, número_de_hojas, metadatos_de_hojas).
    """
    text_parts = []
    sheets_meta = []

    try:
        if PANDAS_AVAILABLE:
            xl = pd.ExcelFile(str(path))
            sheet_names = xl.sheet_names

            for sheet_name in sheet_names:
                try:
                    df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                    if df.empty:
                        continue

                    sheet_text = _dataframe_to_text(df, sheet_name)
                    text_parts.append(sheet_text)

                    sheets_meta.append({
                        'name': sheet_name,
                        'rows': len(df),
                        'cols': len(df.columns),
                        'has_headers': _detect_headers(df),
                    })
                except Exception as e:
                    text_parts.append(f"\n## Hoja: {sheet_name}\n⚠️ Error leyendo hoja: {e}\n")
                    sheets_meta.append({
                        'name': sheet_name,
                        'rows': 0,
                        'cols': 0,
                        'has_headers': False,
                        'error': str(e),
                    })

        elif OPENPYXL_AVAILABLE:
            wb = openpyxl.load_workbook(str(path), data_only=True)
            sheet_names = wb.sheetnames

            for sheet_name in sheet_names:
                try:
                    ws = wb[sheet_name]
                    sheet_text = _worksheet_to_text(ws, sheet_name)
                    text_parts.append(sheet_text)

                    sheets_meta.append({
                        'name': sheet_name,
                        'rows': ws.max_row,
                        'cols': ws.max_column,
                        'has_headers': _detect_headers_openpyxl(ws),
                    })
                except Exception as e:
                    text_parts.append(f"\n## Hoja: {sheet_name}\n⚠️ Error leyendo hoja: {e}\n")
                    sheets_meta.append({
                        'name': sheet_name,
                        'rows': 0,
                        'cols': 0,
                        'has_headers': False,
                        'error': str(e),
                    })

        else:
            return "", 0, []

        full_text = "\n\n".join(text_parts)
        return full_text, len(sheet_names), sheets_meta

    except Exception as e:
        print(f"  ⚠️ Error leyendo Excel {path}: {e}")
        return "", 0, []


def _dataframe_to_text(df, sheet_name: str) -> str:
    """Convierte un DataFrame de pandas a texto Markdown."""
    lines = [f"## Hoja: {sheet_name}", ""]

    has_header = _detect_headers(df)

    if has_header:
        headers = [str(c) if pd.notna(c) else "" for c in df.iloc[0]]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join(["---"] * len(headers)) + "|")
        data_start = 1
    else:
        data_start = 0

    for idx in range(data_start, len(df)):
        row = df.iloc[idx]
        cells = []
        for val in row:
            if pd.isna(val):
                cells.append("")
            else:
                cell_str = str(val).strip().replace("|", r"\|")
                cells.append(cell_str)
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def _worksheet_to_text(ws, sheet_name: str) -> str:
    """Convierte una hoja de openpyxl a texto Markdown."""
    lines = [f"## Hoja: {sheet_name}", ""]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return f"## Hoja: {sheet_name}\n(Vacía)"

    has_header = _detect_headers_openpyxl(ws)

    if has_header:
        headers = [str(c) if c is not None else "" for c in rows[0]]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("|" + "|".join(["---"] * len(headers)) + "|")
        data_rows = rows[1:]
    else:
        data_rows = rows

    for row in data_rows:
        cells = []
        for val in row:
            if val is None:
                cells.append("")
            else:
                cell_str = str(val).strip().replace("|", r"\|")
                cells.append(cell_str)
        lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def _detect_headers(df) -> bool:
    """Detecta si la primera fila de un DataFrame parece ser header."""
    if len(df) < 2:
        return False

    first_row = df.iloc[0]
    second_row = df.iloc[1]

    first_types = [type(v).__name__ for v in first_row if pd.notna(v)]
    second_types = [type(v).__name__ for v in second_row if pd.notna(v)]

    if not first_types or not second_types:
        return False

    first_str_ratio = first_types.count('str') / len(first_types)
    second_str_ratio = second_types.count('str') / len(second_types)

    return first_str_ratio > 0.7 and second_str_ratio < 0.7


def _detect_headers_openpyxl(ws) -> bool:
    """Detecta si la primera fila de una hoja openpyxl parece ser header."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return False

    first_row = [c for c in rows[0] if c is not None]
    second_row = [c for c in rows[1] if c is not None]

    if not first_row or not second_row:
        return False

    first_strs = sum(1 for v in first_row if isinstance(v, str))
    second_strs = sum(1 for v in second_row if isinstance(v, str))

    first_ratio = first_strs / len(first_row)
    second_ratio = second_strs / len(second_row) if second_row else 0

    return first_ratio > 0.7 and second_ratio < 0.7


# ---------------------------------------------------------------------------
# FUNCIONES DE EXTRACCIÓN DE TEXTO (ORIGINALES + NUEVAS)
# ---------------------------------------------------------------------------

def extract_text_from_pdf(path: Path) -> Tuple[str, int]:
    """
    Extrae texto limpio de un archivo PDF.

    Args:
        path: Ruta al archivo PDF.

    Returns:
        Tupla (texto_extraído, número_de_páginas).
    """
    try:
        reader = pypdf.PdfReader(str(path))
        pages_text = []
        for page in reader.pages:
            txt = page.extract_text() or ""
            pages_text.append(txt)
        text = "\n".join(pages_text).strip()
        return text, len(reader.pages)
    except Exception as e:
        print(f"  ⚠️  Error leyendo PDF {path}: {e}")
        return "", 0


def extract_text_from_docx(path: Path) -> Tuple[str, int]:
    """
    Extrae texto de un archivo DOCX.

    Args:
        path: Ruta al archivo DOCX.

    Returns:
        Tupla (texto_extraído, número_de_páginas_estimado).
    """
    try:
        doc = Document(str(path))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs).strip()
        estimated_pages = max(1, len(text) // 3000)
        return text, estimated_pages
    except Exception as e:
        print(f"  ⚠️  Error leyendo DOCX {path}: {e}")
        return "", 0


def extract_text_from_plain(path: Path) -> Tuple[str, int, Optional[List[Dict]]]:
    """
    Lee un archivo de texto plano (Markdown o TXT).
    NUEVO: Detecta y parsea chats de WhatsApp.

    Args:
        path: Ruta al archivo.

    Returns:
        Tupla (texto_extraído, número_de_páginas_estimado, mensajes_chat_o_None).
    """
    try:
        with open(path, 'r', encoding='utf-8') as f:
            text = f.read().strip()

        # Detectar si es chat de WhatsApp
        if is_whatsapp_chat(text, path.name):
            messages = parse_whatsapp_chat(text)
            chat_text = chat_to_text(messages)
            estimated_pages = max(1, len(chat_text) // 3000)
            return chat_text, estimated_pages, messages

        estimated_pages = max(1, len(text) // 3000)
        return text, estimated_pages, None
    except Exception as e:
        print(f"  ⚠️  Error leyendo {path}: {e}")
        return "", 0, None


def extract_text(path: Path) -> Tuple[str, int, Optional[List[Dict]], Optional[List[Dict]]]:
    """
    Dispatcher de extraccion de texto segun extension del archivo.
    NUEVO: Retorna también metadatos de Excel y mensajes de chat.

    Args:
        path: Ruta al archivo.

    Returns:
        Tupla (texto_extraído, número_de_páginas/hojas, mensajes_chat_o_None, sheets_meta_o_None).
    """
    suffix = path.suffix.lower()
    if suffix == '.pdf':
        text, pages = extract_text_from_pdf(path)
        return text, pages, None, None
    elif suffix in ('.docx', '.doc'):
        text, pages = extract_text_from_docx(path)
        return text, pages, None, None
    elif suffix in ('.md', '.txt'):
        text, pages, messages = extract_text_from_plain(path)
        return text, pages, messages, None
    elif suffix in ('.xlsx', '.xls'):
        text, sheets, sheets_meta = extract_text_from_excel(path)
        return text, sheets, None, sheets_meta
    else:
        return "", 0, None, None


# ---------------------------------------------------------------------------
# FUNCIONES DE LIMPIEZA DE TEXTO
# ---------------------------------------------------------------------------

def normalize_newlines(text: str) -> str:
    """Normaliza todos los saltos de linea a \n."""
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text


def fix_ocr_errors(text: str) -> str:
    """Corrige errores comunes de OCR y caracteres especiales."""
    text = text.replace('\ufb01', 'fi')
    text = text.replace('\ufb02', 'fl')
    text = text.replace('\ufb03', 'ffi')
    text = text.replace('\ufb04', 'ffl')
    text = text.replace('\ufb00', 'ff')
    text = re.sub(r'([a-zA-Z])-\n([a-zA-Z])', r'\1\2', text)
    text = re.sub(r'([a-zA-Z])-\r\n?([a-zA-Z])', r'\1\2', text)
    return text


def remove_invisible_chars(text: str) -> str:
    """Elimina caracteres de control e invisibles, preservando acentos."""
    allowed = set('\t\n')
    cleaned = []
    for ch in text:
        cat = unicodedata.category(ch)
        if cat.startswith('C') and ch not in allowed:
            continue
        cleaned.append(ch)
    return ''.join(cleaned)


def collapse_spaces(text: str) -> str:
    """Colapsa espacios multiples en uno solo, preservando indentacion de listas."""
    lines = text.split('\n')
    cleaned_lines = []
    for line in lines:
        stripped = line.lstrip()
        leading = line[:len(line) - len(stripped)]
        stripped = re.sub(r'\s+', ' ', stripped)
        cleaned_lines.append(leading + stripped)
    return '\n'.join(cleaned_lines)


def clean_text(text: str) -> str:
    """
    Pipeline completo de limpieza de texto.
    """
    text = normalize_newlines(text)
    text = fix_ocr_errors(text)
    text = remove_invisible_chars(text)
    text = collapse_spaces(text)
    return text.strip()


# ---------------------------------------------------------------------------
# FUNCIONES DE DETECCION DE METADATOS (MEJORADAS)
# ---------------------------------------------------------------------------

def detect_document_type(text: str, filename: str, is_chat: bool = False,
                         is_excel: bool = False) -> str:
    """
    Detecta el tipo de documento basado en contenido y nombre de archivo.
    NUEVO: Detecta chats y Excel.

    Args:
        text: Contenido del documento.
        filename: Nombre del archivo.
        is_chat: Si es un chat de WhatsApp.
        is_excel: Si es un archivo Excel.

    Returns:
        Tipo de documento en minusculas.
    """
    if is_chat:
        return 'chat'
    if is_excel:
        return 'excel'

    combined = (filename + " " + text[:2000]).lower()

    type_keywords = {
        'acta': ['acta', 'asistentes', 'orden del dia', 'orden del día', 'acuerdos'],
        'relatoria': ['relatoria', 'relator', 'moderador', 'sesion', 'sesión'],
        'manual': ['manual', 'procedimiento', 'instrucciones', 'guia de uso', 'guía de uso'],
        'reglamento': ['reglamento', 'normativa', 'disposiciones', 'articulo', 'artículo'],
        'protocolo': ['protocolo', 'procedimiento estandar', 'estandar de'],
        'carta': ['carta', 'oficio', 'memorando', 'comunicado', 'remitente', 'destinatario'],
        'cronograma': ['cronograma', 'calendario', 'timeline', 'fechas clave', 'hitos'],
        'proyecto': ['proyecto', 'propuesta', 'plan de trabajo', 'plan de accion'],
        'informe': ['informe', 'reporte', 'report', 'analisis', 'análisis', 'estadisticas'],
        'idea': ['idea', 'brainstorm', 'lluvia de ideas', 'concepto inicial'],
        'formato': ['formato', 'formulario', 'plantilla', 'template', 'campos a llenar'],
    }

    scores = {}
    for doc_type, keywords in type_keywords.items():
        score = sum(2 if kw in combined else 0 for kw in keywords)
        scores[doc_type] = score

    fname_lower = filename.lower()
    for doc_type in type_keywords:
        if doc_type in fname_lower:
            scores[doc_type] = scores.get(doc_type, 0) + 5

    best = max(scores, key=scores.get, default='otro')
    if scores.get(best, 0) == 0:
        return 'otro'
    return best


def detect_committee(text: str, filename: str) -> str:
    """
    Detecta el comite al que pertenece el documento.
    """
    combined = (filename + " " + text[:3000]).lower()

    committee_map = {
        'tesoreria': ['tesoreria', 'tesorero', 'presupuesto', 'finanzas', 'viaticos', 'viáticos'],
        'astrolectura': ['astrolectura', 'lectura', 'libro', 'libros'],
        'astroescritura': ['astroescritura', 'escritura', 'redaccion', 'redacción', 'editorial'],
        'etica': ['etica', 'ética', 'codigo de conducta', 'código de conducta', 'comite etico'],
        'charlas': ['charlas', 'conferencia', 'conferencias', 'ponente', 'expositor', 'talk'],
        'administrativo': ['administrativo', 'administracion', 'administración', 'logistica', 'logística'],
        'representacion': ['representacion', 'representación', 'delegado', 'delegados', 'embajador'],
        'cumulo nacional': ['cumulo nacional', 'nacional', 'pais', 'país'],
        'cumulo bogota': ['cumulo bogota', 'cumulo bogotá', 'bogota', 'bogotá', 'local'],
        'observatorio': ['observatorio', 'telescopio', 'observacion', 'observación', 'astronomico'],
        'universidad': ['universidad', 'academico', 'académico', 'facultad', 'carrera'],
        'vip': ['vip', 'invitado especial', 'distinguido'],
        'inventario': ['inventario', 'equipo', 'materiales', 'suministros'],
        'astrochaza': ['astrochaza', 'chaza', 'mercado', 'venta', 'compra'],
        'protocolos': ['protocolo', 'protocolos', 'procedimiento'],
        'astronomia': ['astronomia', 'astronomía', 'astrofotografia', 'astrofotografía'],
    }

    scores = {}
    for committee, keywords in committee_map.items():
        score = sum(2 if kw in combined else 0 for kw in keywords)
        scores[committee] = score

    for committee in committee_map:
        if committee.replace(' ', '') in filename.lower().replace(' ', ''):
            scores[committee] = scores.get(committee, 0) + 5

    best = max(scores, key=scores.get, default='general')
    if scores.get(best, 0) == 0:
        return 'general'
    return best


def detect_scope(text: str, filename: str) -> str:
    """Detecta el alcance geografico o institucional del documento."""
    combined = (filename + " " + text[:2000]).lower()

    scope_map = {
        'cumulo bogota': ['bogota', 'bogotá', 'local', 'sede bogota', 'sede bogotá'],
        'cumulo nacional': ['nacional', 'todo el pais', 'todo el país', 'colombia'],
        'observatorio': ['observatorio', 'observacion', 'observación'],
        'universidad': ['universidad', 'facultad', 'carrera', 'academico', 'académico'],
    }

    for scope, keywords in scope_map.items():
        if any(kw in combined for kw in keywords):
            return scope

    return 'general'


def detect_semester(text: str, filename: str) -> Optional[str]:
    """Detecta el semestre academico al que pertenece el documento."""
    combined = filename + " " + text[:1500]

    match = re.search(r'\b(\d{4})[-\s]?(1|2)\b', combined)
    if match:
        year, sem = match.groups()
        return f"{year}-{sem}"

    dates = extract_dates(text, filename)
    if dates and dates.get('all'):
        first_date = dates['all'][0]
        try:
            dt = datetime.strptime(first_date, '%Y-%m-%d')
            year = dt.year
            semester = 1 if dt.month <= 6 else 2
            return f"{year}-{semester}"
        except ValueError:
            pass

    return None


def detect_status(text: str, filename: str) -> str:
    """Detecta el estado del documento."""
    combined = (filename + " " + text[:2000]).lower()

    if any(kw in combined for kw in ['borrador', 'draft', 'preliminar', 'preliminary']):
        return 'borrador'
    if any(kw in combined for kw in ['propuesta', 'proposed', 'propuesto']):
        return 'propuesta'
    if any(kw in combined for kw in ['historico', 'histórico', 'archivado', 'archived', 'pasado']):
        return 'historico'
    if any(kw in combined for kw in ['vigente', 'actual', 'current', 'en uso', 'aprobado']):
        return 'vigente'

    return 'desconocido'


# ---------------------------------------------------------------------------
# FUNCIONES DE EXTRACCION DE FECHAS
# ---------------------------------------------------------------------------

def _parse_month(month_str: str) -> Optional[int]:
    """Convierte nombre de mes a numero."""
    month_str = month_str.lower().strip()
    return MESES_ES.get(month_str)


def _normalize_date(year: int, month: int, day: int) -> Optional[str]:
    """Valida y normaliza una fecha a ISO."""
    try:
        dt = datetime(year, month, day)
        return dt.strftime('%Y-%m-%d')
    except ValueError:
        return None


def extract_dates(text: str, filename: str) -> Optional[Dict[str, Any]]:
    """
    Extrae todas las fechas del texto y nombre de archivo.
    """
    combined = filename + " " + text[:3000]
    found_dates: Set[str] = set()

    # Patron ISO: 2025-08-15
    for m in re.finditer(r'\b(\d{4})[-/](\d{1,2})[-/](\d{1,2})\b', combined):
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        d = _normalize_date(year, month, day)
        if d:
            found_dates.add(d)

    # Patron espanol corto: 15/08/2025
    for m in re.finditer(r'\b(\d{1,2})[/-](\d{1,2})[/-](\d{4})\b', combined):
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        d = _normalize_date(year, month, day)
        if d:
            found_dates.add(d)

    # Patron espanol largo: 15 de agosto de 2025
    pattern_long = r'\b(\d{1,2})\s+de\s+([a-záéíóúñ]+)\s+de(?:l)?\s+(\d{4})\b'
    for m in re.finditer(pattern_long, combined, re.IGNORECASE):
        day = int(m.group(1))
        month = _parse_month(m.group(2))
        year = int(m.group(3))
        if month:
            d = _normalize_date(year, month, day)
            if d:
                found_dates.add(d)

    # Rangos: 27-29 de enero de 2026
    pattern_range = r'\b(\d{1,2})\s*[-–—]\s*(\d{1,2})\s+de\s+([a-záéíóúñ]+)\s+de(?:l)?\s+(\d{4})\b'
    for m in re.finditer(pattern_range, combined, re.IGNORECASE):
        start_day = int(m.group(1))
        end_day = int(m.group(2))
        month = _parse_month(m.group(3))
        year = int(m.group(4))
        if month:
            for day in range(start_day, end_day + 1):
                d = _normalize_date(year, month, day)
                if d:
                    found_dates.add(d)

    # Multiples: 12, 13 y 15 de agosto de 2025
    pattern_multi = r'\b(\d{1,2})(?:,\s+(\d{1,2}))(?:\s+y\s+(\d{1,2}))?\s+de\s+([a-záéíóúñ]+)\s+de(?:l)?\s+(\d{4})\b'
    for m in re.finditer(pattern_multi, combined, re.IGNORECASE):
        days = [int(m.group(1))]
        if m.group(2):
            days.append(int(m.group(2)))
        if m.group(3):
            days.append(int(m.group(3)))
        month = _parse_month(m.group(4))
        year = int(m.group(5))
        if month:
            for day in days:
                d = _normalize_date(year, month, day)
                if d:
                    found_dates.add(d)

    if not found_dates:
        return None

    sorted_dates = sorted(found_dates)
    return {
        'start': sorted_dates[0],
        'end': sorted_dates[-1],
        'all': sorted_dates,
    }


# ---------------------------------------------------------------------------
# FUNCIONES DE EXTRACCION DE PERSONAS Y FIRMANTES
# ---------------------------------------------------------------------------

def extract_people(text: str) -> List[Dict[str, str]]:
    """
    Extrae nombres de personas y sus posibles cargos.
    """
    people = []
    seen = set()

    lines = text.split('\n')
    for line in lines:
        line_lower = line.lower()
        if not any(role in line_lower for role in ROLES):
            continue

        name_matches = re.findall(
            r'\b([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+){1,3})\b',
            line
        )

        for name in name_matches:
            name = name.strip()
            if len(name) < 5:
                continue
            detected_role = 'asistente'
            for role in ROLES:
                if role in line_lower:
                    detected_role = role
                    break

            key = f"{name.lower()}|{detected_role}"
            if key not in seen:
                seen.add(key)
                people.append({'name': name, 'role': detected_role})

    return people


def extract_signatures(text: str) -> List[Dict[str, str]]:
    """
    Extrae firmas del documento.
    """
    signatures = []
    seen = set()

    signature_blocks = re.split(
        r'(?:firma[s]?|firmado[s]?|firmantes|firmado por)[:\s]*',
        text,
        flags=re.IGNORECASE
    )

    for block in signature_blocks[1:]:
        block_text = block[:500]
        lines = block_text.split('\n')[:10]

        for line in lines:
            line = line.strip()
            if not line or len(line) < 3:
                continue

            name_matches = re.findall(
                r'\b([A-ZÁÉÍÓÚÑ][a-záéíóúñ]+(?:\s+[A-ZÁÉÍÓÚÑ][a-záéíóúñ]+){1,3})\b',
                line
            )

            for name in name_matches:
                name = name.strip()
                if len(name) < 5:
                    continue

                detected_role = 'firmante'
                line_lower = line.lower()
                for role in ROLES:
                    if role in line_lower:
                        detected_role = role
                        break

                key = f"{name.lower()}|{detected_role}"
                if key not in seen:
                    seen.add(key)
                    signatures.append({'name': name, 'role': detected_role})

    return signatures


# ---------------------------------------------------------------------------
# FUNCIONES DE EXTRACCION DE SECCIONES
# ---------------------------------------------------------------------------

def extract_sections(text: str) -> Dict[str, str]:
    """
    Separa el documento en secciones identificadas por encabezados.
    """
    sections = {}

    header_pattern = '|'.join(re.escape(h) for h in SECTION_HEADERS)
    regex = re.compile(
        rf'(?:^|\n)\s*(?:#{{1,4}}\s*)?\b({header_pattern})\b[:\.\s]*(?:\n|$)',
        re.IGNORECASE
    )

    matches = list(regex.finditer(text))

    if not matches:
        sections['contenido'] = text.strip()
        return sections

    for i, match in enumerate(matches):
        header = match.group(1).lower().strip()
        start = match.end()
        end = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        content = text[start:end].strip()
        if content:
            sections[header] = content

    if matches[0].start() > 0:
        preamble = text[:matches[0].start()].strip()
        if preamble:
            sections['contexto'] = preamble

    return sections


# ---------------------------------------------------------------------------
# FUNCIONES DE GENERACION DE RESUMEN Y PALABRAS CLAVE
# ---------------------------------------------------------------------------

def generate_summary(text: str, sections: Dict[str, str]) -> str:
    """
    Genera un resumen automatico sin usar IA.
    """
    parts = []

    first_para = text.split('\n\n')[0].strip() if text else ""
    if len(first_para) > 50:
        parts.append(first_para)

    if 'contexto' in sections and sections['contexto']:
        parts.append(sections['contexto'][:500])

    if 'objetivos' in sections and sections['objetivos']:
        parts.append("Objetivos: " + sections['objetivos'][:500])

    if 'objetivo' in sections and sections['objetivo']:
        parts.append("Objetivo: " + sections['objetivo'][:500])

    summary = " ".join(parts)

    words = summary.split()
    if len(words) > 300:
        summary = " ".join(words[:300]) + "..."

    return summary.strip()


def generate_keywords(text: str) -> List[str]:
    """
    Genera palabras clave eliminando stopwords.
    """
    words = re.findall(r'\b[a-záéíóúñ]{3,}\b', text.lower())
    filtered = [w for w in words if w not in STOPWORDS_ES]
    counter = Counter(filtered)
    keywords = [word for word, count in counter.most_common(20)]
    return keywords


# ---------------------------------------------------------------------------
# FUNCIONES DE CHUNKING INTELIGENTE (MEJORADO)
# ---------------------------------------------------------------------------

def smart_chunk(text: str, sections: Dict[str, str],
                chat_messages: Optional[List[Dict]] = None) -> List[Dict[str, Any]]:
    """
    Divide el texto en chunks inteligentes respetando secciones y estructura.
    NUEVO: Si es un chat, usa chunking especializado por mensajes.
    """
    if chat_messages:
        return chunk_chat_messages(chat_messages, "")

    chunks = []
    chunk_id = 0

    for section_name, section_text in sections.items():
        if not section_text.strip():
            continue

        words = section_text.split()

        if len(words) <= CHUNK_MAX_WORDS:
            chunks.append({
                'id': f"chunk_{chunk_id:05d}",
                'section': section_name,
                'text': section_text.strip(),
                'tokens': len(words),
                'chars': len(section_text),
            })
            chunk_id += 1
        else:
            paragraphs = section_text.split('\n\n')
            current_chunk = []
            current_words = 0

            for para in paragraphs:
                para = para.strip()
                if not para:
                    continue

                para_words = len(para.split())

                if current_words + para_words > CHUNK_MAX_WORDS and current_chunk:
                    chunk_text = '\n\n'.join(current_chunk)
                    chunks.append({
                        'id': f"chunk_{chunk_id:05d}",
                        'section': section_name,
                        'text': chunk_text.strip(),
                        'tokens': current_words,
                        'chars': len(chunk_text),
                    })
                    chunk_id += 1
                    current_chunk = [para]
                    current_words = para_words
                else:
                    current_chunk.append(para)
                    current_words += para_words

            if current_chunk:
                chunk_text = '\n\n'.join(current_chunk)
                chunks.append({
                    'id': f"chunk_{chunk_id:05d}",
                    'section': section_name,
                    'text': chunk_text.strip(),
                    'tokens': current_words,
                    'chars': len(chunk_text),
                })
                chunk_id += 1

    return chunks


# ---------------------------------------------------------------------------
# FUNCIONES DE CONTEO Y ESTADISTICAS
# ---------------------------------------------------------------------------

def count_attendees(people: List[Dict[str, str]]) -> int:
    """Cuenta el numero de asistentes/participantes."""
    return len(people)


def count_agreements(text: str) -> int:
    """Cuenta acuerdos mencionados en el texto."""
    patterns = [
        r'\bacuerdo[s]?\b',
        r'\bse acord[oó]\b',
        r'\bqued[oó] establecido\b',
        r'\bresoluci[oó]n\b',
    ]
    count = 0
    for pattern in patterns:
        count += len(re.findall(pattern, text, re.IGNORECASE))
    return count


def count_tasks(text: str) -> int:
    """Cuenta tareas mencionadas en el texto."""
    patterns = [
        r'\btarea[s]?\b',
        r'\bpending[s]?\b',
        r'\basignar\b',
        r'\basignado\b',
        r'\bresponsable\b',
        r'\bentrega\b',
        r'\bplazo\b',
    ]
    count = 0
    for pattern in patterns:
        count += len(re.findall(pattern, text, re.IGNORECASE))
    return count


def count_signatures(signatures: List[Dict[str, str]]) -> int:
    """Cuenta firmas extraidas."""
    return len(signatures)


def compute_file_hash(path: Path) -> str:
    """Calcula SHA-256 del archivo."""
    sha256 = hashlib.sha256()
    try:
        with open(path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b""):
                sha256.update(chunk)
        return sha256.hexdigest()
    except Exception as e:
        print(f"  ⚠️  Error calculando hash de {path}: {e}")
        return ""


def build_stats(text: str, people: List[Dict], signatures: List[Dict],
                chunks: List[Dict], pages: int,
                chat_messages: Optional[List[Dict]] = None,
                sheets_meta: Optional[List[Dict]] = None) -> Dict[str, Any]:
    """
    Construye el bloque de estadisticas del documento.
    NUEVO: Incluye stats de chats y Excel.
    """
    words = len(text.split())
    chars = len(text)

    stats = {
        'pages': pages,
        'words': words,
        'characters': chars,
        'attendees': count_attendees(people),
        'agreements': count_agreements(text),
        'tasks': count_tasks(text),
        'signatures': count_signatures(signatures),
        'chunks': len(chunks),
    }

    if chat_messages:
        stats['chat'] = {
            'total_messages': len(chat_messages),
            'unique_participants': len(set(m['sender'] for m in chat_messages)),
            'participants': sorted(set(m['sender'] for m in chat_messages)),
            'media_omitted': sum(1 for m in chat_messages if m['type'] == 'media_omitted'),
            'links': sum(1 for m in chat_messages if m['type'] == 'link'),
        }

    if sheets_meta:
        stats['excel'] = {
            'total_sheets': len(sheets_meta),
            'sheets': sheets_meta,
        }

    return stats


# ---------------------------------------------------------------------------
# FUNCIONES DE CONSTRUCCION DE DOCUMENTO
# ---------------------------------------------------------------------------

def generate_document_id(doc_type: str, committee: str, dates: Optional[Dict],
                         filename: str, index: int) -> str:
    """
    Genera un identificador unico permanente para el documento.
    """
    type_abbr = doc_type.upper()[:4]
    comm_abbr = committee.upper().replace(' ', '')[:6]

    year = "0000"
    if dates and dates.get('start'):
        year = dates['start'][:4]
    else:
        m = re.search(r'\b(\d{4})\b', filename)
        if m:
            year = m.group(1)

    return f"{type_abbr}-{comm_abbr}-{year}-{index:03d}"


def build_document(path: Path, index: int) -> Optional[Dict[str, Any]]:
    """
    Construye el objeto documento completo a partir de un archivo.
    NUEVO: Maneja chats de WhatsApp y Excel.
    """
    print(f"  📄 {path.name}", end=" ... ")

    # 1. Extraer texto
    text, pages, chat_messages, sheets_meta = extract_text(path)
    if not text:
        print("⚠️  (vacio)")
        return None

    is_chat = chat_messages is not None
    is_excel = sheets_meta is not None

    # 2. Limpiar texto
    text = clean_text(text)

    # 3. Detectar metadatos
    doc_type = detect_document_type(text, path.name, is_chat=is_chat, is_excel=is_excel)
    committee = detect_committee(text, path.name)
    scope = detect_scope(text, path.name)
    semester = detect_semester(text, path.name)
    status = detect_status(text, path.name)
    dates = extract_dates(text, path.name)

    # 4. Extraer personas y firmas
    if is_chat:
        people = [{'name': sender, 'role': 'participante'}
                  for sender in sorted(set(m['sender'] for m in chat_messages))]
        signatures = []
    else:
        people = extract_people(text)
        signatures = extract_signatures(text)

    # 5. Extraer secciones
    sections = extract_sections(text)

    # 6. Generar resumen y keywords
    summary = generate_summary(text, sections)
    keywords = generate_keywords(text)

    # 7. Chunks inteligentes
    chunks = smart_chunk(text, sections, chat_messages=chat_messages)

    # 8. Estadisticas
    stats = build_stats(text, people, signatures, chunks, pages,
                        chat_messages=chat_messages, sheets_meta=sheets_meta)

    # 9. Hash del archivo
    file_hash = compute_file_hash(path)

    # 10. Informacion del archivo
    file_info = {
        'name': path.name,
        'path': str(path),
        'extension': path.suffix.lower().lstrip('.'),
        'size_bytes': path.stat().st_size,
        'modified_at': datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
        'extracted_at': datetime.now().isoformat(),
        'hash_sha256': file_hash,
    }

    # 11. Generar ID
    doc_id = generate_document_id(doc_type, committee, dates, path.name, index)

    # 12. Asignar IDs de documento a chunks
    for chunk in chunks:
        chunk['document_id'] = doc_id

    # 13. Exportar chat como JSON independiente
    chat_export_path = None
    if is_chat and CHAT_EXPORT_LINES:
        chat_export_path = export_chat_json(chat_messages, doc_id, path.name)

    # 14. Relaciones
    relations = {
        'same_committee': [],
        'consecutive': [],
        'versions': [],
    }

    # 15. Version
    version = "1.0"
    v_match = re.search(r'v(\d+(?:\.\d+)?)', path.name, re.IGNORECASE)
    if v_match:
        version = v_match.group(1)

    # Metadatos específicos
    extra_meta = {}
    if is_chat and chat_messages:
        extra_meta['chat'] = {
            'total_messages': len(chat_messages),
            'participants': sorted(set(m['sender'] for m in chat_messages)),
            'date_range': {
                'first': chat_messages[0]['date'] if chat_messages else None,
                'last': chat_messages[-1]['date'] if chat_messages else None,
            },
            'exported_to': str(chat_export_path) if chat_export_path else None,
        }
    if is_excel:
        extra_meta['excel'] = {
            'sheets': sheets_meta,
        }

    print(f"✓ {doc_id} ({stats['words']} palabras, {len(chunks)} chunks)", end="")
    if is_chat:
        print(f" [💬 {len(chat_messages)} msgs]", end="")
    if is_excel:
        print(f" [📊 {len(sheets_meta)} hojas]", end="")
    print()

    doc = {
        'document_id': doc_id,
        'document_type': doc_type,
        'committee': committee,
        'scope': scope,
        'semester': semester,
        'status': status,
        'version': version,
        'dates': dates,
        'participants': people,
        'signatures': signatures,
        'summary': summary,
        'keywords': keywords,
        'sections': sections,
        'chunks': chunks,
        'stats': stats,
        'file': file_info,
        'relations': relations,
    }

    if extra_meta:
        doc['extra'] = extra_meta

    return doc


# ---------------------------------------------------------------------------
# FUNCIONES DE RELACIONES ENTRE DOCUMENTOS
# ---------------------------------------------------------------------------

def build_relations(documents: List[Dict[str, Any]]) -> None:
    """
    Detecta y establece relaciones entre documentos.

    Relaciones:
        - same_committee: documentos del mismo comite.
        - consecutive: actas con IDs consecutivos.
        - versions: documentos con mismo tipo+comite+anio pero diferente version.

    Args:
        documents: Lista de documentos (modificada in-place).
    """
    # Indice por comite
    by_committee = defaultdict(list)
    for doc in documents:
        by_committee[doc['committee']].append(doc['document_id'])

    # Indice por tipo+comite+anio
    by_group = defaultdict(list)
    for doc in documents:
        key = f"{doc['document_type']}-{doc['committee']}"
        if doc['dates'] and doc['dates'].get('start'):
            key += f"-{doc['dates']['start'][:4]}"
        by_group[key].append(doc)

    for doc in documents:
        doc_id = doc['document_id']

        # Mismo comite
        same = [d for d in by_committee[doc['committee']] if d != doc_id]
        doc['relations']['same_committee'] = same[:10]  # Limitar a 10

        # Consecutivos (mismo prefijo, numero +1 o -1)
        prefix = '-'.join(doc_id.split('-')[:-1])
        try:
            num = int(doc_id.split('-')[-1])
            for other in documents:
                if other['document_id'] == doc_id:
                    continue
                other_prefix = '-'.join(other['document_id'].split('-')[:-1])
                try:
                    other_num = int(other['document_id'].split('-')[-1])
                    if other_prefix == prefix and abs(other_num - num) == 1:
                        doc['relations']['consecutive'].append(other['document_id'])
                except ValueError:
                    continue
        except ValueError:
            pass

        # Versiones
        group_key = f"{doc['document_type']}-{doc['committee']}"
        if doc['dates'] and doc['dates'].get('start'):
            group_key += f"-{doc['dates']['start'][:4]}"
        versions = [d['document_id'] for d in by_group[group_key] if d['document_id'] != doc_id]
        doc['relations']['versions'] = versions[:5]


# ---------------------------------------------------------------------------
# FUNCIONES DE CONSTRUCCION DE METADATA GLOBAL
# ---------------------------------------------------------------------------

def build_metadata(documents: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Construye el indice global de metadatos (kb-metadata.json).
    NUEVO: Incluye stats de chats y Excel.

    Args:
        documents: Lista de todos los documentos procesados.

    Returns:
        Diccionario de metadata global.
    """
    # Conteos por tipo
    type_counts = Counter(d['document_type'] for d in documents)

    # Conteos por comite
    committee_counts = Counter(d['committee'] for d in documents)

    # Conteos por anio
    year_counts = Counter()
    for d in documents:
        if d['dates'] and d['dates'].get('start'):
            year_counts[d['dates']['start'][:4]] += 1

    # Conteos por semestre
    semester_counts = Counter()
    for d in documents:
        if d['semester']:
            semester_counts[d['semester']] += 1

    # Conteos por estado
    status_counts = Counter(d['status'] for d in documents)

    # Conteos por carpeta
    folder_counts = Counter()
    for d in documents:
        folder = d['file']['path'].split(os.sep)[1] if os.sep in d['file']['path'] else 'root'
        folder_counts[folder] += 1

    # Personas detectadas (unicas)
    all_people = set()
    for d in documents:
        for p in d['participants']:
            all_people.add(p['name'])

    # Firmantes (unicos)
    all_signers = set()
    for d in documents:
        for s in d['signatures']:
            all_signers.add(s['name'])

    # Fechas minima y maxima
    all_dates = []
    for d in documents:
        if d['dates'] and d['dates'].get('all'):
            all_dates.extend(d['dates']['all'])

    date_range = {
        'earliest': min(all_dates) if all_dates else None,
        'latest': max(all_dates) if all_dates else None,
    }

    docs_without_date = sum(1 for d in documents if d['dates'] is None)

    # Palabras clave mas frecuentes
    all_keywords = []
    for d in documents:
        all_keywords.extend(d['keywords'])
    top_keywords = [kw for kw, _ in Counter(all_keywords).most_common(30)]

    # Documentos con mas tareas
    docs_by_tasks = sorted(
        documents,
        key=lambda d: d['stats']['tasks'],
        reverse=True
    )[:10]
    top_task_docs = [d['document_id'] for d in docs_by_tasks]

    # Total de chunks
    total_chunks = sum(d['stats']['chunks'] for d in documents)
    total_tokens = sum(
        sum(c['tokens'] for c in d['chunks'])
        for d in documents
    )

    # NUEVO: Stats de chats
    chat_docs = [d for d in documents if d['document_type'] == 'chat']
    total_chat_messages = sum(d['stats'].get('chat', {}).get('total_messages', 0) for d in chat_docs)
    all_chat_participants = set()
    for d in chat_docs:
        for p in d['stats'].get('chat', {}).get('participants', []):
            all_chat_participants.add(p)

    # NUEVO: Stats de Excel
    excel_docs = [d for d in documents if d['document_type'] == 'excel']
    total_excel_sheets = sum(d['stats'].get('excel', {}).get('total_sheets', 0) for d in excel_docs)

    # NUEVO: Chats exportados
    exported_chats = []
    for d in chat_docs:
        if d.get('extra', {}).get('chat', {}).get('exported_to'):
            exported_chats.append({
                'document_id': d['document_id'],
                'file': d['file']['name'],
                'path': d['extra']['chat']['exported_to'],
            })

    return {
        'generated_at': datetime.now().isoformat(),
        'total_documents': len(documents),
        'total_chunks': total_chunks,
        'total_tokens': total_tokens,
        'counts': {
            'by_type': dict(type_counts),
            'by_committee': dict(committee_counts),
            'by_year': dict(year_counts),
            'by_semester': dict(semester_counts),
            'by_status': dict(status_counts),
            'by_folder': dict(folder_counts),
        },
        'people': {
            'all_detected': sorted(all_people),
            'count': len(all_people),
            'signers': sorted(all_signers),
            'signers_count': len(all_signers),
        },
        'dates': {
            'range': date_range,
            'without_date': docs_without_date,
        },
        'keywords': {
            'top_30': top_keywords,
        },
        'rankings': {
            'most_tasks': top_task_docs,
        },
        'documents': [d['document_id'] for d in documents],
        # NUEVO: Secciones de chat y Excel
        'chats': {
            'total_chat_documents': len(chat_docs),
            'total_messages': total_chat_messages,
            'unique_participants': len(all_chat_participants),
            'all_participants': sorted(all_chat_participants),
            'exported_files': exported_chats,
        },
        'excel': {
            'total_excel_documents': len(excel_docs),
            'total_sheets': total_excel_sheets,
        },
    }


# ---------------------------------------------------------------------------
# FUNCIONES DE EXPORTACION
# ---------------------------------------------------------------------------

def export_json(documents: List[Dict[str, Any]], metadata: Dict[str, Any]) -> None:
    """
    Exporta los archivos JSON finales.
    NUEVO: Tambien exporta indice de chats.

    Args:
        documents: Lista de documentos procesados.
        metadata: Metadata global.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # kb.json
    kb_path = OUTPUT_DIR / "kb.json"
    with open(kb_path, 'w', encoding='utf-8') as f:
        json.dump(documents, f, ensure_ascii=False, indent=2)
    kb_size = kb_path.stat().st_size / 1024

    # kb-metadata.json
    meta_path = OUTPUT_DIR / "kb-metadata.json"
    with open(meta_path, 'w', encoding='utf-8') as f:
        json.dump(metadata, f, ensure_ascii=False, indent=2)
    meta_size = meta_path.stat().st_size / 1024

    # NUEVO: Indice de chats exportados
    if metadata['chats']['exported_files']:
        chats_index_path = CHATS_DIR / "_index.json"
        with open(chats_index_path, 'w', encoding='utf-8') as f:
            json.dump(metadata['chats']['exported_files'], f, ensure_ascii=False, indent=2)

    print(f"\n  💾 kb.json          → {kb_path} ({kb_size:.1f} KB)")
    print(f"  💾 kb-metadata.json → {meta_path} ({meta_size:.1f} KB)")
    if metadata['chats']['exported_files']:
        print(f"  💾 chats/           → {CHATS_DIR}/ ({len(metadata['chats']['exported_files'])} archivos)")


# ---------------------------------------------------------------------------
# FUNCION PRINCIPAL
# ---------------------------------------------------------------------------

def main() -> None:
    """Punto de entrada principal del generador de base de conocimiento."""
    print("=" * 70)
    print("🔨  Cúmulo Knowledge Base Generator v2.0")
    print("     + Soporte Excel (.xlsx, .xls)")
    print("     + Soporte Chats WhatsApp (.txt)")
    print("     + Exportacion de chats a JSON independientes")
    print("=" * 70)

    if not INPUT_DIR.exists():
        print(f"\n❌ No se encontro la carpeta '{INPUT_DIR}' en {Path.cwd()}")
        print("   Crea una carpeta llamada 'documentos' con tus archivos.")
        exit(1)

    # Recolectar archivos soportados (NUEVO: .xlsx, .xls)
    supported_exts = {'.pdf', '.docx', '.doc', '.md', '.txt', '.xlsx', '.xls'}
    files = sorted([
        p for p in INPUT_DIR.rglob('*')
        if p.is_file() and p.suffix.lower() in supported_exts and not p.name.startswith('.')
    ])

    if not files:
        print(f"\n❌ No se encontraron documentos en '{INPUT_DIR}'")
        exit(1)

    print(f"\n📂 {len(files)} documentos encontrados:")
    for f in files:
        icon = "📄"
        if f.suffix.lower() in ('.xlsx', '.xls'):
            icon = "📊"
        elif f.suffix.lower() == '.txt':
            icon = "💬"
        print(f"   {icon} {f.relative_to(INPUT_DIR)}")
    print()

    documents = []
    for idx, path in enumerate(files, start=1):
        doc = build_document(path, idx)
        if doc:
            documents.append(doc)

    if not documents:
        print("\n❌ No se pudo procesar ningun documento.")
        exit(1)

    print(f"\n✅ {len(documents)} documentos procesados exitosamente.")

    # Construir relaciones entre documentos
    print("\n🔗 Analizando relaciones entre documentos...")
    build_relations(documents)

    # Construir metadata global
    print("📊 Generando indice global...")
    metadata = build_metadata(documents)

    # Exportar
    print("\n💾 Exportando archivos JSON...")
    export_json(documents, metadata)

    # Resumen final
    print("\n" + "=" * 70)
    print("📋 RESUMEN FINAL")
    print("=" * 70)
    print(f"  Documentos procesados: {metadata['total_documents']}")
    print(f"  Chunks generados:      {metadata['total_chunks']}")
    print(f"  Tokens totales:        {metadata['total_tokens']:,}")
    print(f"  Personas detectadas:   {metadata['people']['count']}")
    print(f"  Firmantes:             {metadata['people']['signers_count']}")
    print(f"  Rango de fechas:       {metadata['dates']['range']['earliest']} → {metadata['dates']['range']['latest']}")
    print(f"  Sin fecha:             {metadata['dates']['without_date']}")

    # NUEVO: Resumen de chats
    if metadata['chats']['total_chat_documents'] > 0:
        print(f"\n  💬 Chats WhatsApp:")
        print(f"     Documentos:  {metadata['chats']['total_chat_documents']}")
        print(f"     Mensajes:    {metadata['chats']['total_messages']:,}")
        print(f"     Participantes: {metadata['chats']['unique_participants']}")
        print(f"     Exportados:  {len(metadata['chats']['exported_files'])} archivos en docs/chats/")

    # NUEVO: Resumen de Excel
    if metadata['excel']['total_excel_documents'] > 0:
        print(f"\n  📊 Archivos Excel:")
        print(f"     Documentos: {metadata['excel']['total_excel_documents']}")
        print(f"     Hojas:      {metadata['excel']['total_sheets']}")

    print("\n  Por tipo:")
    for doc_type, count in metadata['counts']['by_type'].items():
        print(f"    • {doc_type}: {count}")
    print("\n  Por comite:")
    for committee, count in metadata['counts']['by_committee'].items():
        print(f"    • {committee}: {count}")
    print("\n" + "=" * 70)
    print("🚀 Proximos pasos:")
    print("   1. Sube docs/kb.json y docs/kb-metadata.json a GitHub")
    print("   2. Sube docs/chats/*.json si quieres acceso directo a chats")
    print("   3. Activa GitHub Pages: Settings → Pages → /docs folder")
    print("   4. Tu KB estara en: https://tu-usuario.github.io/cumulo")
    print("=" * 70)


if __name__ == "__main__":
    main()
